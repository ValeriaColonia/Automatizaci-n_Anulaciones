import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from PIL import Image as PILImage
import os
from datetime import datetime

# Configuración de directorios
TEMPLATE_DIR = "templates"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.join(TEMPLATE_DIR, "img_proc"), exist_ok=True)

SEDES = ["VIVA 1A IPS ESPINAL", "VIVA 1A IPS GUAMO", "VIVA 1A IPS MARIQUITA"]
SOLICITANTES_DEV = ["Diana Sanchez", "Pedro Cabezas", "Mónica Sandoval"]
MOTIVOS = ["Demanda Inducida", "Paciente PYP", "Paciente Gestante"]
SOLICITANTES_AUX = ["Yenny Mendez", "Valentina Triana", "Carmen Sanchez", "Yazmin Camargo", "Yeimy Cardenas", "Valeria Colonia"]

# Configuración de la página y diseño elegante
st.set_page_config(
    page_title="Sistema de Anulaciones - VIVA 1A",
    page_icon="🏥",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Estilos CSS personalizados para una interfaz sofisticada y elegante
st.markdown("""
    <style>
    .main {
        background-color: #f8f9fa;
    }
    .stApp {
        max-width: 800px;
        margin: 0 auto;
    }
    h1 {
        color: #1A365D;
        font-family: 'Helvetica Neue', sans-serif;
        font-weight: 700;
        text-align: center;
        padding-bottom: 10px;
        border-bottom: 2px solid #E2E8F0;
    }
    h3 {
        color: #2B6CB0;
        font-family: 'Helvetica Neue', sans-serif;
    }
    .stButton>button {
        width: 100%;
        border-radius: 8px;
        font-weight: 600;
        padding: 0.6rem 1rem;
        transition: all 0.3s ease;
    }
    .stButton>button[kind="primary"] {
        background-color: #2B6CB0;
        color: white;
        border: none;
    }
    .stButton>button[kind="primary"]:hover {
        background-color: #1A365D;
    }
    .footer {
        position: fixed;
        left: 0;
        bottom: 0;
        width: 100%;
        background-color: #ffffff;
        color: #718096;
        text-align: center;
        padding: 10px;
        font-size: 13px;
        border-top: 1px solid #E2E8F0;
        z-index: 999;
    }
    </style>
""", unsafe_allow_html=True)

def verificar_reinicio_diario():
    consolidado_path = os.path.join(OUTPUT_DIR, "consolidado_anulaciones_dia.xlsx")
    hoy = datetime.now().strftime("%Y-%m-%d")
    
    if os.path.exists(consolidado_path):
        try:
            wb = openpyxl.load_workbook(consolidado_path)
            ws = wb.active
            primera_fecha = ws.cell(row=4, column=1).value
            if primera_fecha and primera_fecha != hoy:
                os.remove(consolidado_path)
        except Exception:
            pass

def preparar_logo(ws):
    for img in ws._images:
        try:
            img_pil = PILImage.open(img.ref)
            if img_pil.mode in ('RGBA', 'LA') or (img_pil.mode == 'P' and 'transparency' in img_pil.info):
                img_pil = img_pil.convert('RGBA')
                bg = PILImage.new("RGB", img_pil.size, (255, 255, 255))
                bg.paste(img_pil, mask=img_pil.split()[3])
                processed_path = os.path.join(TEMPLATE_DIR, "img_proc", "logo_limpio.png")
                bg.save(processed_path, "PNG")
                img.ref = processed_path
            else:
                img_pil.convert('RGB').save(os.path.join(TEMPLATE_DIR, "img_proc", "logo_limpio.png"), "PNG")
                img.ref = os.path.join(TEMPLATE_DIR, "img_proc", "logo_limpio.png")
        except Exception:
            pass

def generar_devolucion(datos):
    template_path = os.path.join(TEMPLATE_DIR, "Formato-Devolucion-Cuota-Moderadora.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    preparar_logo(ws)
    
    ahora = datetime.now()
    ws['D7'] = ahora.strftime("%d")
    ws['E7'] = ahora.strftime("%m")
    ws['F7'] = ahora.strftime("%Y")
    
    ws['B10'] = datos['sede']
    ws['E10'] = datos['solicitante_dev']
    ws['B13'] = datos['doc_paciente']
    ws['E13'] = datos['nom_paciente']
    ws['B16'] = datos['valor']
    
    cell_concepto = ws['E16']
    cell_concepto.value = "CTA MODER"
    cell_concepto.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    ws['B19'] = "Motivo de Devolución"
    
    cell_motivo = ws['B20']
    cell_motivo.value = datos['motivo']
    cell_motivo.alignment = Alignment(horizontal='center', vertical='center')

    ws['B29'] = datos['solicitante_aux']
    ws['B30'] = "Nombre del Solicitante"

    output_path = os.path.join(OUTPUT_DIR, f"Devolucion_{datos['doc_paciente']}.xlsx")
    wb.save(output_path)
    return output_path

def agregar_al_consolidado(datos):
    verificar_reinicio_diario()
    consolidado_path = os.path.join(OUTPUT_DIR, "consolidado_anulaciones_dia.xlsx")
    
    if os.path.exists(consolidado_path):
        wb = openpyxl.load_workbook(consolidado_path)
        ws = wb.active
    else:
        template_path = os.path.join(TEMPLATE_DIR, "Formato para solicitud de anulaciones.xlsx")
        wb = openpyxl.load_workbook(template_path)
        ws = wb["Formato Solicitud anulacion"] if "Formato Solicitud anulacion" in wb.sheetnames else wb.active

    fila = 4
    while ws.cell(row=fila, column=1).value is not None:
        fila += 1

    ws.cell(row=fila, column=1, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=fila, column=2, value=datos['sede'])
    ws.cell(row=fila, column=3, value=datos['doc_paciente'])
    ws.cell(row=fila, column=4, value=datos['nom_paciente'])
    ws.cell(row=fila, column=5, value=datos['valor'])
    ws.cell(row=fila, column=6, value=datos['concepto'])
    ws.cell(row=fila, column=7, value=datos['motivo'])
    ws.cell(row=fila, column=8, value=datos.get('observaciones', ''))
    ws.cell(row=fila, column=9, value=datos.get('cobro', 'Sí'))
    ws.cell(row=fila, column=10, value=datos['solicitante_aux'])
    ws.cell(row=fila, column=11, value=datos.get('soporte', 'Sí'))

    wb.save(consolidado_path)
    return consolidado_path, fila - 3

# Encabezado visual
st.title("🏥 Gestión de Anulaciones y Copagos")
st.markdown("<p style='text-align: center; color: #4A5568;'>Plataforma automatizada para control diario de formatos y devoluciones</p>", unsafe_allow_html=True)

verificar_reinicio_diario()
consolidado_path = os.path.join(OUTPUT_DIR, "consolidado_anulaciones_dia.xlsx")

# Lectura y conteo por sedes
conteo_hoy = 0
conteos_sedes = {sede: 0 for sede in SEDES}

if os.path.exists(consolidado_path):
    wb_temp = openpyxl.load_workbook(consolidado_path)
    ws_temp = wb_temp.active
    f = 4
    while ws_temp.cell(row=f, column=1).value is not None:
        conteo_hoy += 1
        s = ws_temp.cell(row=f, column=2).value
        if s in conteos_sedes:
            conteos_sedes[s] += 1
        f += 1

# Métricas elegantes por sede
st.markdown("### 📊 Resumen Diario de Anulaciones")
col_m1, col_m2, col_m3, col_m4 = st.columns(4)
with col_m1:
    st.metric(label="Total General", value=conteo_hoy)
with col_m2:
    st.metric(label="Espinal", value=conteos_sedes["VIVA 1A IPS ESPINAL"])
with col_m3:
    st.metric(label="Guamo", value=conteos_sedes["VIVA 1A IPS GUAMO"])
with col_m4:
    st.metric(label="Mariquita", value=conteos_sedes["VIVA 1A IPS MARIQUITA"])

st.markdown("---")

with st.form("form_anulacion", clear_on_submit=False):
    st.subheader("📝 Registrar Nueva Devolución")
    
    col1, col2 = st.columns(2)
    with col1:
        sede = st.selectbox("Sede y Ciudad", SEDES)
        solicitante_dev = st.selectbox("Solicita la devolución", SOLICITANTES_DEV)
        doc_paciente = st.text_input("Documento del Paciente")
        nom_paciente = st.text_input("Nombre del Paciente")
        valor = st.number_input("Valor a devolver / Factura", min_value=0.0, step=100.0)
    
    with col2:
        motivo = st.selectbox("Motivo de anulación", MOTIVOS)
        solicitante_aux = st.selectbox("Auxiliar de Admisiones", SOLICITANTES_AUX)
        cobro = st.selectbox("¿Se realizó cobro?", ["Sí", "No"])
        soporte = st.selectbox("¿Se anexa soporte?", ["Sí", "No"])
        observaciones = st.text_input("Observaciones (opcional)")
    
    st.markdown("<br>", unsafe_allow_html=True)
    enviar = st.form_submit_button("✨ Generar Formato y Registrar en Consolidado", type="primary")

if enviar:
    if not doc_paciente or not nom_paciente:
        st.error("⚠️ Por favor complete el documento y nombre del paciente.")
    else:
        datos = {
            'sede': sede,
            'solicitante_dev': solicitante_dev,
            'doc_paciente': doc_paciente,
            'nom_paciente': nom_paciente,
            'valor': valor,
            'concepto': "CTA MODER",
            'motivo': motivo,
            'solicitante_aux': solicitante_aux,
            'observaciones': observaciones,
            'cobro': cobro,
            'soporte': soporte
        }
        
        path_individual = generar_devolucion(datos)
        path_consolidado, total = agregar_al_consolidado(datos)
        
        st.success(f"¡Proceso completado con éxito! Total de registros hoy: {total}")
        st.session_state['last_individual'] = path_individual
        st.session_state['last_consolidado'] = path_consolidado
        st.rerun()

st.markdown("---")
st.subheader("📥 Zona de Descargas")

col_d1, col_d2 = st.columns(2)

with col_d1:
    if 'last_individual' in st.session_state and os.path.exists(st.session_state['last_individual']):
        with open(st.session_state['last_individual'], "rb") as file:
            st.download_button(
                label="📄 Descargar Devolución Individual",
                data=file,
                file_name=os.path.basename(st.session_state['last_individual']),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("Genere una anulación primero.")

with col_d2:
    if os.path.exists(consolidado_path):
        # Selector de sede para descarga filtrada
        sede_seleccionada_descarga = st.selectbox("Filtrar consolidado por Sede:", ["Todas las Sedes"] + SEDES)
        
        if sede_seleccionada_descarga == "Todas las Sedes":
            with open(consolidado_path, "rb") as file:
                st.download_button(
                    label="📊 Descargar Consolidado General",
                    data=file,
                    file_name="consolidado_anulaciones_dia.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            # Crear un archivo filtrado temporalmente para la sede seleccionada
            wb_orig = openpyxl.load_workbook(consolidado_path)
            ws_orig = wb_orig.active
            
            template_path = os.path.join(TEMPLATE_DIR, "Formato para solicitud de anulaciones.xlsx")
            wb_filtrado = openpyxl.load_workbook(template_path)
            ws_filtrado = wb_filtrado["Formato Solicitud anulacion"] if "Formato Solicitud anulacion" in wb_filtrado.sheetnames else wb_filtrado.active
            
            # Limpiar filas existentes en el template filtrado desde la fila 4
            f_clean = 4
            while ws_filtrado.cell(row=f_clean, column=1).value is not None:
                for c in range(1, 12):
                    ws_filtrado.cell(row=f_clean, column=c, value=None)
                f_clean += 1
                
            f_orig = 4
            f_dest = 4
            while ws_orig.cell(row=f_orig, column=1).value is not None:
                if ws_orig.cell(row=f_orig, column=2).value == sede_seleccionada_descarga:
                    for c in range(1, 12):
                        ws_filtrado.cell(row=f_dest, column=c, value=ws_orig.cell(row=f_orig, column=c).value)
                    f_dest += 1
                f_orig += 1
                
            temp_filtrado_path = os.path.join(OUTPUT_DIR, f"consolidado_{sede_seleccionada_descarga.replace(' ', '_')}.xlsx")
            wb_filtrado.save(temp_filtrado_path)
            
            with open(temp_filtrado_path, "rb") as file:
                st.download_button(
                    label=f"📊 Descargar Consolidado ({sede_seleccionada_descarga.split()[-1]})",
                    data=file,
                    file_name=f"consolidado_{sede_seleccionada_descarga.replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.warning("Sin registros consolidados hoy.")

# Sección de administración / limpieza diaria manual
st.markdown("---")
with st.expander("⚙️ Opciones de Administración (Reinicio Diario)"):
    st.write("Si necesitas reiniciar los contadores y limpiar el consolidado actual de forma manual para comenzar un nuevo turno o día, haz clic en el siguiente botón:")
    if st.button("🗑️ Limpiar y Reiniciar Día (Borrar Consolidado)", type="secondary"):
        if os.path.exists(consolidado_path):
            os.remove(consolidado_path)
        if 'last_consolidado' in st.session_state:
            del st.session_state['last_consolidado']
        if 'last_individual' in st.session_state:
            del st.session_state['last_individual']
        st.success("¡Consolidado y contadores reiniciados a cero exitosamente!")
        st.rerun()

# Pie de página fijo corporativo con tu crédito
st.markdown("""
    <div class="footer">
        Realizado por: <b>Valeria Colonia</b> &nbsp;|&nbsp; VIVA 1A IPS 
    </div>
""", unsafe_allow_html=True)