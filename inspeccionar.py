import openpyxl
import os

print("--- INSPECCIONANDO PLANTILLAS ---")
for f in os.listdir("templates"):
    if f.endswith(".xlsx"):
        path = os.path.join("templates", f)
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"\nArchivo: {f} | Hojas: {wb.sheetnames}")
        s = wb.active
        for row in list(s.iter_rows(values_only=False))[:15]:
            for c in row:
                if c.value is not None:
                    print(f"  [{c.coordinate}]: {c.value}")